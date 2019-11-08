import getopt
import os
import sys
import time
from DBconnection.dbconnection import dbconnection
from util.util import *
from Constants import *
import logging
import xlsxwriter
from openpyxl import *
# from openpyxl import Workbook
# from openpyxl import load_workbook
import Constants


def getMergefilelist(single_xls_path,key):

    mtime = lambda f: os.stat(os.path.join(single_xls_path, f)).st_mtime
    listfile=list(sorted(os.listdir(single_xls_path), key=mtime))
    # listfile=list(os.listdir(single_xls_path))
    print "listfile=",listfile
    localfile = [f for f in listfile if isfile(join(single_xls_path, f))]
    # print "localfile=",localfile
    mergefilelist = []
    print "key=",key
    for item in localfile:
        if item.__contains__(key):
            mergefilelist.append(item)
    print "mergefilelist=",mergefilelist
    return mergefilelist

def mergefile(key):
    cwd = os.getcwd()
    single_xls_path = os.path.join(cwd, Constants.taskresult_folder)
    mergefilefolder=os.path.join(single_xls_path,Constants.mergeOriginal_folder)
    # key = "customQueue";
    mergefilelist=getMergefilelist(mergefilefolder,key)
    print mergefilelist
    filenum=mergefilelist.__len__()
    print filenum
    excelcontentlist=[]
    for item in mergefilelist:
        filename=os.path.join(mergefilefolder,item)
        excelcontentlist.append(readexcel(filename))
    # print excelcontentlist
    newfilename=key+".xlsx"
    filepath = os.path.join(os.path.join(cwd, taskresult_folder),merge_folder)
    newfile = os.path.join(filepath, newfilename)
    try:
        workbook = xlsxwriter.Workbook(newfile)
        worksheet = workbook.add_worksheet()
        worksheet.set_column(0, 0, 13)
        worksheet.set_column(1, 1, 20)
        worksheet.set_column(2, 2, 13)
        worksheet.set_column(3, 3, 15)
        worksheet.set_column(4, 4, 60)
        worksheet.set_column(5, 5, 10)
        for i in range(0, excelcontentlist.__len__() - 1):
            worksheet.set_column(6 + i, 6 + i, 10)
        format_head = workbook.add_format({'bold': True})
        format_head.set_align('center')
        format_head.set_valign("vcenter")
        format_head.set_text_wrap()
        format_head.set_border()

        format_body = workbook.add_format({'bold': False})
        format_body.set_align('center')
        format_body.set_valign("vcenter")
        format_body.set_text_wrap()
        format_body.set_border()

        format_pass = workbook.add_format({'bold': False})
        format_pass.set_align('center')
        format_pass.set_valign("vcenter")
        format_pass.set_text_wrap()
        format_pass.set_border()
        format_pass.set_bg_color('green')

        format_fail = workbook.add_format({'bold': False})
        format_fail.set_align('center')
        format_fail.set_valign("vcenter")
        format_fail.set_text_wrap()
        format_fail.set_border()
        format_fail.set_bg_color('red')
        row = 0;
        for i in range(0, excelcontentlist.__len__()):
            print "i=",i
            row = 0
            count = 0
            if i == 0:
                for item in excelcontentlist[0]:
                    if row == 0:
                        worksheet.write(0, 0, cons_test_case_id, format_head)
                        worksheet.write(0, 1, cons_test_case_name, format_head)
                        worksheet.write(0, 2, cons_test_task_id, format_head)
                        worksheet.write(0, 3, cons_test_instance_id, format_head)
                        worksheet.write(0, 4, cons_test_instance_name, format_head)
                        worksheet.write(0, 5, cons_test_status, format_head)
                    else:
                        worksheet.write(row, 0, item[0], format_body)
                        worksheet.write(row, 1, item[1], format_body)
                        worksheet.write(row, 2, item[2], format_body)
                        worksheet.write(row, 3, item[3], format_body)
                        worksheet.write(row, 4, item[4], format_body)
                        if item[5] == 'pass':
                            worksheet.write(row, 5, item[5], format_pass)
                        else:
                            worksheet.write(row, 5, item[5], format_fail)
                    row += 1
            else:
                for j in range(0,len(excelcontentlist[0])):
                    if row == 0:
                        worksheet.write(0, 5 + i, cons_test_status, format_head)
                    else:
                        # print "row=", row
                        # print "excelcontentlist[i][row][3]=", excelcontentlist[i][row-count][3]
                        # print "excelcontentlist[0][row][3]=", excelcontentlist[0][row][3]
                        # print "row-count=",row-count
                        # print "----",len(excelcontentlist[i])
                        if row-count+1>len(excelcontentlist[i]):
                            count += row-count+1-len(excelcontentlist[i])
                        if excelcontentlist[i][row-count][3] == excelcontentlist[0][row][3]:
                            if excelcontentlist[i][row-count][5] == 'pass':
                                worksheet.write(row, 5 + i, excelcontentlist[i][row-count][5], format_pass)
                            else:
                                worksheet.write(row, 5 + i, excelcontentlist[i][row-count][5], format_fail)
                        else:
                            count +=1

                            print "Don't have this row!"
                        # print "count=", count
                    row += 1
        workbook.close()
    except Exception as e:
        print "Generate excel report failed. Exception -" + str(e)

def readexcel(filename):
    wb= load_workbook(filename)
    sheet = wb['Sheet1']
    row_count= sheet.max_row
    col_count=sheet.max_column
    excelcontent=[]
    for i in range(0,row_count):
        listrow=[]
        for j in range(0,col_count):
            listrow.append(sheet.cell(row=i+1,column=j+1).value)
        excelcontent.append(listrow)
    return excelcontent

def Usage():
    print 'Generate new task Usage is as folllowing:\n'
    print '-h,--help: print help message.\n'
    print '-k --key : file name key word \n'
    print '---------exemple-------------------\n'

    print 'python mergeResult.py --key inboundpolicy_ThirdPart  \n'


def main(argv):
    # LOG_FILE = 'getcaseResult.log'
    # logging.basicConfig(format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S%p', filename=LOG_FILE,level=logging.DEBUG)
    key=""
    #print "----argv----",argv
    if len(argv)<3:
        Usage()
        sys.exit(2)
    try:
        options, args = getopt.getopt(argv[1:], 'k:h', ["key=", "--help"])
    except getopt.GetoptError, err:
        print str(err)
        Usage()
        sys.exit(2)
    for opt, arg in options:
        if opt in ('-h', '--help'):
            Usage()
            sys.exit(0)
        elif opt in ('-k', '--key'):
            key = arg
        else:
            print 'unhandled option'
            Usage()
            sys.exit(2)
    keylist=[]
    if key.__contains__(","):
        keylist=key.split(",")
    else:
        keylist.append(key)
    for item in keylist:
        mergefile(item)


if __name__ == '__main__':
    main(sys.argv)