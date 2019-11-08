import getopt
import os
import sys
import time
from DBconnection.dbconnection import dbconnection
from util.util import *
from Constants import *
import logging
import xlsxwriter
from openpyxl import *
# from openpyxl import Workbook
# from openpyxl import load_workbook
import Constants


def getfilelist(single_xls_path):

    ctime = lambda f: os.stat(os.path.join(single_xls_path, f)).st_ctime
    listfile=list(sorted(os.listdir(single_xls_path), key=ctime))
    # listfile=list(os.listdir(single_xls_path))
    # print "listfile=",listfile
    localfile = [f for f in listfile if isfile(join(single_xls_path, f))]
    # print localfile
    return localfile

def getResult(single_xls_path):
    # cwd = os.getcwd()
    # single_xls_path = "E:\\robot_bq\\LegacyAutomation\\taskResult\\8_5_round2"

    # key = "customQueue";
    filelist=getfilelist(single_xls_path)
    # print filelist
    filenum=filelist.__len__()
    print filenum
    totalexcellist=[]
    for item in filelist:
        filename=os.path.join(single_xls_path,item)
        totalexcellist.append(readexcel(filename))
    # print excelcontentlist
    # print totalexcellist
    totalcasenum=0
    totalpassnum=0
    totalfailnum=0
    failitemsingelinfo=[]
    failitem=[]
    notrunitemsingleinfo=[]
    notrunitem=[]
    for item in totalexcellist:
        if item[3]>0:
            failitemsingelinfo.append(item[0])
            failitemsingelinfo.append(item[3])
        if item[1]==0:
            notrunitemsingleinfo.append(item[0])
            notrunitemsingleinfo.append(item[1])
            notrunitemsingleinfo.append(item[2])
            notrunitemsingleinfo.append(item[3])
        totalcasenum += item[1]
        totalpassnum += item[2]
        totalfailnum += item[3]

    failitem.append(failitemsingelinfo)
    notrunitem.append(notrunitemsingleinfo)
    totalinfolist=[]
    totalinfolist.append(totalcasenum)
    totalinfolist.append(totalpassnum)
    totalinfolist.append(totalfailnum)
    print "======Total case number:",totalcasenum
    print "======Total pass case number:", totalpassnum
    print "======Total failed case number:", totalfailnum
    print "=======Failed items :==================="
    print failitem
    print "=======Not run items:======================"
    print notrunitem

    hungitem=[]

    for item in totalexcellist:
        if 'action_language' in item[0] and item[1] != 18:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(18)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'AddressRewriting' in item[0] and item[1] != 54:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(54)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'alert_event' in item[0] and item[1] != 5:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(5)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'archiveQueue' in item[0] and item[1] != 20:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(20)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'AV' in item[0] and item[1] != 128:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(128)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'AS' in item[0] and item[1] != 66:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(66)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'BlockMessages' in item[0] and item[1] != 29:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(29)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'ConnectControl' in item[0] and item[1] != 16:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(16)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'ContentFilter' in item[0] and 'ContentFilter_Language' not in item[0] and item[1] != 250:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(250)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'ContentFilter_Language' in item[0] and item[1] != 20:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(20)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'customQueue' in item[0] and item[1] != 21:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(21)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'decryptionfailQueue' in item[0] and item[1] != 18:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(18)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'Directory_Attacks' in item[0] and item[1] != 5:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(5)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'Disclaimer' in item[0] and item[1] != 40:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(40)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'DomainGroup' in item[0] and item[1] != 16:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(16)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'encryptionfailQueue' in item[0] and item[1] != 18:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(18)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'EnforceTLS' in item[0] and item[1] != 22:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(22)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'inboundOutbound_general' in item[0] and item[1] != 4:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(4)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'inboundpolicy_Firstpart' in item[0] and item[1] != 418:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(418)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'inboundpolicy_SecondPart' in item[0] and item[1] != 189:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(189)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'inboundpolicy_ThirdPart' in item[0] and item[1] != 34:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(34)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'internalPolicy_Firstpart' in item[0] and item[1] != 420:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(420)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'internalPolicy_SecondPart' in item[0] and item[1] != 195:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(195)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'internalPolicy_ThirdPart' in item[0] and item[1] != 56:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(56)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'IPGroup' in item[0] and item[1] != 3:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(3)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'Logs_general_audit' in item[0] and item[1] != 50:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(50)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'mailRouting' in item[0] and item[1] != 28:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(28)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'messageControl' in item[0] and item[1] != 12:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(12)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'MessageQueues' in item[0] and item[1] != 25:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(25)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'outboundPolicy_Firstpart' in item[0] and item[1] != 399:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(399)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'outboundPolicy_SecondPart' in item[0] and item[1] != 190:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(190)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'outboundPolicy_ThirdPart' in item[0] and item[1] != 42:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(42)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'PEMActivities' in item[0] and item[1] != 62:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(62)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'RelayControl' in item[0] and item[1] != 19:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(19)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'Reporting' in item[0] and item[1] != 13:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(13)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'spamQueue' in item[0] and item[1] != 31:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(31)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'TrafficShaping' in item[0] and item[1] != 108:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(108)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'TrueSourceIP' in item[0] and item[1] != 40:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(40)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'UA' in item[0] and item[1] != 3:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(3)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'UserAuthentication' in item[0] and item[1] != 17:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(17)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)
        if 'virusQueue' in item[0] and item[1] != 23:
            hungitemsingleinfo = []
            hungitemsingleinfo.append(item[0])
            hungitemsingleinfo.append(23)
            hungitemsingleinfo.append(item[1])
            hungitem.append(hungitemsingleinfo)

    # hungitem.append(hungitemsingleinfo)
    print "============Hung Item==============="
    # print hungitem
    for item in hungitem:
        print "******Hung case name:",item[0],"  Total case number should be:",item[1]," Only run case number:",item[2]


def readexcel(filename):
    statuslist=[]
    wb= load_workbook(filename)
    sheet = wb['Sheet1']
    row_count= sheet.max_row
    col_count=sheet.max_column
    statuslist.append(filename)
    statuslist.append(row_count-1)
    # excelcontent=[]
    passnum=0
    failednum=0
    for i in range(0,row_count):
        listrow=[]
        if sheet.cell(row=i+1,column=6).value == "pass":
            passnum +=1
        elif sheet.cell(row=i+1,column=6).value == "test_status":
            pass
        else:
            # print "---",sheet.cell(row=i+1,column=6).value
            failednum += 1
    statuslist.append(passnum)
    statuslist.append(failednum)
    return statuslist

# mergefile()

def Usage():
    print 'Get the pass and failed case number as folllowing:\n'
    print '-h,--help: print help message.\n'
    print '-p --path : xls file path  \n'
    print '---------exemple-------------------\n'

    print 'python CaculateCaseNumber.py --path   \n'


def main(argv):
    # LOG_FILE = 'getcaseResult.log'
    # logging.basicConfig(format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S%p', filename=LOG_FILE,level=logging.DEBUG)
    path=""
    #print "----argv----",argv
    if len(argv)<3:
        Usage()
        sys.exit(2)
    try:
        options, args = getopt.getopt(argv[1:], 'p:h', ["path=", "--help"])
    except getopt.GetoptError, err:
        print str(err)
        Usage()
        sys.exit(2)
    for opt, arg in options:
        if opt in ('-h', '--help'):
            Usage()
            sys.exit(0)
        elif opt in ('-p', '--path'):
            path = arg
        else:
            print 'unhandled option'
            Usage()
            sys.exit(2)
    getResult(path)


if __name__ == '__main__':
    main(sys.argv)


